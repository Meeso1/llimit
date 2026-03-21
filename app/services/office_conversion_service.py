import io

import docx
import openpyxl

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


class OfficeConversionService:
    """Service for converting Office documents to plain text."""

    def extract_text(self, content_type: str, content: bytes) -> str:
        """Extract plain text from an Office document."""
        if content_type == XLSX_CONTENT_TYPE:
            return self._xlsx_to_text(content)
        elif content_type == DOCX_CONTENT_TYPE:
            return self._docx_to_text(content)
        raise ValueError(f"Not a supported Office content type: {content_type}")

    def _xlsx_to_text(self, content: bytes) -> str:
        """Convert xlsx binary content to a plain-text representation of all sheets."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = [
                "\t".join("" if cell.value is None else str(cell.value) for cell in row)
                for row in ws.iter_rows()
            ]
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
        return "\n\n".join(parts)

    def _docx_to_text(self, content: bytes) -> str:
        """Convert docx binary content to plain text, preserving paragraphs and tables."""
        document = docx.Document(io.BytesIO(content))
        parts: list[str] = []
        for block in document.element.body:
            tag = block.tag.split("}")[-1]
            if tag == "p":
                text = "".join(node.text or "" for node in block.iter() if node.text)
                if text.strip():
                    parts.append(text)
            elif tag == "tbl":
                table = next((t for t in document.tables if t._tbl is block), None)
                if table:
                    for row in table.rows:
                        parts.append("\t".join(cell.text for cell in row.cells))
        return "\n".join(parts)
